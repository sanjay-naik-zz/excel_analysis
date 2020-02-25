import re
import string
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def cell_to_col_row(cell):
    # type: (object) -> object
    # Using re.compile() + re.match() + re.groups()
    # Splitting text and number in string
    cell = cell.replace('$', '')
    print cell
    temp = re.compile("([a-zA-Z]+)([0-9]+)")
    res = temp.match(cell).groups()
    column = res[0]
    row = res[1]
    return column, row


def row_col_to_cell(row,column):
    # type: (str, str) -> str
    return column_num_to_string(column)+str(row)


def column_num_to_string(n):
    # type: (number_string) -> string
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


def column_string_to_num(col):
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num


def read_excel(filename):
    workbook = load_workbook(filename=filename)
    sheet_dict = {}
    for sheet_name in workbook.sheetnames:
        sheet_dict[sheet_name] = workbook[sheet_name]
    return sheet_dict


def neighbour_cells(cell):
    column, row = cell_to_col_row(cell)
    column_num = column_string_to_num(column)
    row_num = int(row)
    next_row_cell = row_col_to_cell(row_num + 1, column_num)
    next_col_cell = row_col_to_cell(row_num, column_num + 1)
    prev_row_cell = None
    prev_col_cell = None
    if row_num > 1:
        prev_row_cell = row_col_to_cell(row_num - 1, column_num)
    if column_num>1:
        prev_col_cell = row_col_to_cell(row_num + 1, column_num)

    return next_row_cell, next_col_cell, prev_row_cell, prev_col_cell

def neighbour_cell_values(cell, sheet_to_analyze, **sheet_dist):
    next_row_cell, next_col_cell, prev_row_cell, prev_col_cell = neighbour_cells(cell)
    next_row_cell_value = sheet_dist[sheet_to_analyze][next_row_cell].value
    next_col_cell_value = sheet_dist[sheet_to_analyze][next_col_cell].value

    prev_row_cell_value = None
    prev_col_cell_value = None
    if prev_row_cell is not None:
        prev_row_cell_value = sheet_dist[sheet_to_analyze][prev_row_cell].value
    if prev_col_cell is not None:
        prev_col_cell_value = sheet_dist[sheet_to_analyze][prev_col_cell].value
    return next_row_cell_value, next_col_cell_value, prev_row_cell_value, prev_col_cell_value
