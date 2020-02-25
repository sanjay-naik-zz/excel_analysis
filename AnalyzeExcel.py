from openpyxl import load_workbook
import re
from parser import *
from excel_utils import *

if __name__ == '__main__':
    # Read Multi Sheet Excel into A dictionary of workbooks
    sheet_dist = read_excel('/Users/sanjay/Desktop/NF-SA.xlsx')

    # Picking Sheet to be Analyzed for Analysis

    sheet_to_analyze = 'SA-Ratios'
    sheet = sheet_dist[sheet_to_analyze]

    col_size = sheet.max_column
    row_size = sheet.max_row

    print col_size, row_size

    parsed_excel_info_dict = {}
    formulas = {}

    for column in range(1, 4):  # iterating through all the columns
        for row in range(1, 76):  # Iterating through all the rows
            current_cell = row_col_to_cell(row, column)
            next_row_cell, next_col_cell, prev_row_cell, prev_col_cell = neighbour_cells(current_cell)

            # Finding the value of all the neighbouring cells

            current_cell_value = sheet[current_cell].value
            next_row_cell_value, next_col_cell_value, prev_row_cell_value, prev_col_cell_value = neighbour_cell_values(
                current_cell, sheet_to_analyze, **sheet_dist)

            # print next_row_cell_value, next_col_cell_value, prev_row_cell_value, prev_col_cell_value

            if current_cell_value is not None and isinstance(current_cell_value,
                                                             str):  # Considering only the cells where formulas exist
                if current_cell_value[0] == '=':
                    # Check if the Cell/Expression is already parsed.
                    if parsed_excel_info_dict.get(current_cell) is None:
                        parsed_excel_info_dict[current_cell] = get_parsed_info_dict(current_cell_value[1:])

                        # Check if the expression is formula and not reference
                        # And Check if the previous column value is not calculated
                        if (prev_col_cell_value is None or prev_col_cell_value[
                            0] != '=') and parsed_excel_info_dict.get(current_cell).get("operators_size") > 0:
                            # Parse next column if not Parsed Already
                            if parsed_excel_info_dict.get(next_col_cell) is None:
                                parsed_excel_info_dict[next_col_cell] = get_parsed_info_dict(next_col_cell_value[1:])
                            current_cell_info = parsed_excel_info_dict.get(current_cell)
                            next_col_cell_info = parsed_excel_info_dict.get(next_col_cell)
                            if current_cell_info.get("operands_size") == next_col_cell_info.get(
                                    "operands_size") and current_cell_info.get(
                                "operators_size") == next_col_cell_info.get(
                                "operators_size"):
                                initial_column = column
                                initial_column_value = sheet[row_col_to_cell(row, initial_column)].value
                                while initial_column_value is None or initial_column_value[0] == '=':
                                    initial_column = initial_column - 1
                                    # print row, initial_column, initial_column_value, column, parsed_excel_info_dict.get(current_cell).get("operands")
                                    initial_column_value = sheet[row_col_to_cell(row, initial_column)].value
                                column_definition = initial_column_value
                                iteration = 0
                                operand_definition = str()
                                for operand in current_cell_info.get("operands"):
                                    braces = set('()')
                                    quotes = set('"\'')
                                    quotes_open = False
                                    sheet_name = str()
                                    lookup_cell = str()
                                    for character in operand:
                                        if character in braces:
                                            operand_definition = operand_definition + character
                                        elif character in quotes:
                                            if not quotes_open:
                                                quotes_open = True
                                            else:
                                                quotes_open = False
                                        elif quotes_open:
                                            sheet_name = sheet_name + character
                                        else:
                                            lookup_cell = lookup_cell + character
                                    if len(sheet_name) == 0:
                                        print current_cell_info
                                        print lookup_cell, current_cell_info.get("operands"), current_cell_info.get("operators"), current_cell, current_cell_value
                                        if re.search('[a-zA-Z]', lookup_cell):
                                            operand_definition = operand_definition + lookup_cell_definition(lookup_cell, sheet_to_analyze, **sheet_dist)
                                        else:
                                            operand_definition = operand_definition + lookup_cell
                                    if len(current_cell_info.get("operators")) > iteration:
                                        operand_definition = operand_definition + " " + \
                                                             current_cell_info.get("operators")[iteration] + " "
                                        iteration = iteration + 1
                                formulas[column_definition] = operand_definition
    print formulas
